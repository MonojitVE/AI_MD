"""Report generation service — builds PDF and Excel exports for dashboard data.

Keeps report *formatting* separate from data fetching. Routers fetch data
using the existing analytics/defects/maintenance query functions and pass
plain dicts/lists in here to be rendered.
"""

import io
from datetime import datetime
from typing import Dict, List, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
)


BRAND_COLOR = colors.HexColor("#1E3A5F")


# ─── Excel Helpers ─────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(*(Side(style="thin", color="D0D0D0"),) * 4)


def _autofit_columns(ws, rows: List[List[Any]]):
    """Roughly autosize columns based on content length."""
    widths = {}
    for row in rows:
        for i, val in enumerate(row):
            length = len(str(val)) if val is not None else 0
            widths[i] = max(widths.get(i, 10), min(length + 2, 50))
    for i, width in widths.items():
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def _write_sheet(ws, title: str, headers: List[str], rows: List[List[Any]]):
    ws.title = title[:31]  # Excel sheet name limit
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    _autofit_columns(ws, [headers] + rows)


def build_excel_workbook(sheets: Dict[str, Dict[str, Any]]) -> io.BytesIO:
    """
    Build a multi-sheet Excel workbook.

    sheets: { "Sheet Name": {"headers": [...], "rows": [[...], ...]} }
    """
    wb = Workbook()
    wb.remove(wb.active)

    for name, content in sheets.items():
        ws = wb.create_sheet()
        _write_sheet(ws, name, content["headers"], content["rows"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── PDF Helpers ───────────────────────────────────────────────────

def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ReportTitle", fontSize=20, textColor=BRAND_COLOR,
        spaceAfter=4, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        name="ReportSubtitle", fontSize=10, textColor=colors.grey,
        spaceAfter=20,
    ))
    styles.add(ParagraphStyle(
        name="SectionHeader", fontSize=13, textColor=BRAND_COLOR,
        spaceBefore=16, spaceAfter=8, fontName="Helvetica-Bold",
    ))
    return styles


def _data_table(headers: List[str], rows: List[List[Any]], col_widths: Optional[List[float]] = None) -> Table:
    table_data = [headers] + rows
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return t


def _build_pdf(title: str, subtitle: str, story_sections: List) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = _pdf_styles()
    story = [
        Paragraph(title, styles["ReportTitle"]),
        Paragraph(subtitle, styles["ReportSubtitle"]),
    ]
    story.extend(story_sections)
    doc.build(story)
    buf.seek(0)
    return buf


def _timestamp() -> str:
    return f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"


# ─── Maintenance Cost Report ────────────────────────────────────────

def maintenance_pdf(data: Dict) -> io.BytesIO:
    styles = _pdf_styles()
    story = []

    story.append(Paragraph(f"Total Maintenance Cost: ${data['total_cost']:,.2f}", styles["SectionHeader"]))

    story.append(Paragraph("Cost by Equipment (Top 10)", styles["SectionHeader"]))
    rows = [[k, f"${v:,.2f}"] for k, v in data["by_equipment"].items()]
    story.append(_data_table(["Equipment", "Cost"], rows, col_widths=[3.5 * inch, 2 * inch]))

    story.append(Paragraph("Cost by Maintenance Type", styles["SectionHeader"]))
    rows = [[k.title(), f"${v:,.2f}"] for k, v in data["by_type"].items()]
    story.append(_data_table(["Type", "Cost"], rows, col_widths=[3.5 * inch, 2 * inch]))

    story.append(Paragraph("Monthly Trend", styles["SectionHeader"]))
    rows = [[m["month"], m["count"], f"${m['cost']:,.2f}"] for m in data["monthly_trend"]]
    story.append(_data_table(["Month", "Jobs", "Cost"], rows, col_widths=[2 * inch, 1.5 * inch, 2 * inch]))

    return _build_pdf("Maintenance Cost Report", _timestamp(), story)


def maintenance_excel(data: Dict) -> io.BytesIO:
    sheets = {
        "By Equipment": {
            "headers": ["Equipment", "Cost ($)"],
            "rows": [[k, round(v, 2)] for k, v in data["by_equipment"].items()],
        },
        "By Type": {
            "headers": ["Maintenance Type", "Cost ($)"],
            "rows": [[k.title(), round(v, 2)] for k, v in data["by_type"].items()],
        },
        "Monthly Trend": {
            "headers": ["Month", "Job Count", "Cost ($)"],
            "rows": [[m["month"], m["count"], round(m["cost"], 2)] for m in data["monthly_trend"]],
        },
    }
    return build_excel_workbook(sheets)


# ─── Defect Report ───────────────────────────────────────────────────

def defects_pdf(stats: Dict, defect_list: List[Dict]) -> io.BytesIO:
    styles = _pdf_styles()
    story = []

    story.append(Paragraph("Summary", styles["SectionHeader"]))
    summary_rows = [
        ["Total Defects", stats["total_defects"]],
        ["Critical", stats["critical_count"]],
        ["High", stats["high_count"]],
        ["Medium", stats["medium_count"]],
        ["Low", stats["low_count"]],
        ["Resolved", stats["resolved_count"]],
        ["Unresolved", stats["unresolved_count"]],
        ["Resolution Rate", f"{stats['resolution_rate']}%"],
    ]
    story.append(_data_table(["Metric", "Value"], summary_rows, col_widths=[3 * inch, 2.5 * inch]))

    story.append(Paragraph("By Defect Type", styles["SectionHeader"]))
    rows = [[k.replace("_", " ").title(), v] for k, v in stats["by_type"].items()]
    story.append(_data_table(["Defect Type", "Count"], rows, col_widths=[3.5 * inch, 2 * inch]))

    story.append(Paragraph("Top Equipment by Defect Count", styles["SectionHeader"]))
    rows = [[k, v] for k, v in stats["by_equipment"].items()]
    story.append(_data_table(["Equipment", "Count"], rows, col_widths=[3.5 * inch, 2 * inch]))

    if defect_list:
        story.append(PageBreak())
        story.append(Paragraph("Recent Defect Records", styles["SectionHeader"]))
        rows = [
            [
                d["timestamp"][:10],
                d["equipment_name"],
                d["defect_type"].replace("_", " ").title(),
                d["severity"].title(),
                f"{d['confidence_score'] * 100:.0f}%",
                "Yes" if d["is_resolved"] else "No",
            ]
            for d in defect_list[:40]
        ]
        story.append(_data_table(
            ["Date", "Equipment", "Type", "Severity", "Confidence", "Resolved"],
            rows,
            col_widths=[0.9 * inch, 1.6 * inch, 1.3 * inch, 0.9 * inch, 0.9 * inch, 0.8 * inch],
        ))

    return _build_pdf("Defect Detection Report", _timestamp(), story)


def defects_excel(stats: Dict, defect_list: List[Dict]) -> io.BytesIO:
    sheets = {
        "Summary": {
            "headers": ["Metric", "Value"],
            "rows": [
                ["Total Defects", stats["total_defects"]],
                ["Critical", stats["critical_count"]],
                ["High", stats["high_count"]],
                ["Medium", stats["medium_count"]],
                ["Low", stats["low_count"]],
                ["Resolved", stats["resolved_count"]],
                ["Unresolved", stats["unresolved_count"]],
                ["Resolution Rate (%)", stats["resolution_rate"]],
            ],
        },
        "By Type": {
            "headers": ["Defect Type", "Count"],
            "rows": [[k.replace("_", " ").title(), v] for k, v in stats["by_type"].items()],
        },
        "By Equipment": {
            "headers": ["Equipment", "Count"],
            "rows": [[k, v] for k, v in stats["by_equipment"].items()],
        },
        "Records": {
            "headers": ["Date", "Equipment", "Type", "Severity", "Confidence (%)", "Resolved", "Description"],
            "rows": [
                [
                    d["timestamp"][:10],
                    d["equipment_name"],
                    d["defect_type"].replace("_", " ").title(),
                    d["severity"].title(),
                    round(d["confidence_score"] * 100, 1),
                    "Yes" if d["is_resolved"] else "No",
                    d.get("description") or "",
                ]
                for d in defect_list
            ],
        },
    }
    return build_excel_workbook(sheets)


# ─── OEE Report ───────────────────────────────────────────────────────

def oee_pdf(oee: Dict, equipment_health: List[Dict]) -> io.BytesIO:
    styles = _pdf_styles()
    story = []

    story.append(Paragraph("Overall Equipment Effectiveness", styles["SectionHeader"]))
    rows = [
        ["Overall OEE", f"{oee['overall']}%"],
        ["Availability", f"{oee['availability']}%"],
        ["Performance", f"{oee['performance']}%"],
        ["Quality", f"{oee['quality']}%"],
    ]
    story.append(_data_table(["Metric", "Value"], rows, col_widths=[3 * inch, 2.5 * inch]))

    if equipment_health:
        story.append(Paragraph("Equipment Health Breakdown", styles["SectionHeader"]))
        rows = [[e["name"], e["type"], f"{e['health_score']}", e["status"].title()] for e in equipment_health]
        story.append(_data_table(
            ["Equipment", "Type", "Health Score", "Status"],
            rows,
            col_widths=[2 * inch, 1.8 * inch, 1.2 * inch, 1.2 * inch],
        ))

    return _build_pdf("OEE Report", _timestamp(), story)


def oee_excel(oee: Dict, equipment_health: List[Dict]) -> io.BytesIO:
    sheets = {
        "OEE Summary": {
            "headers": ["Metric", "Value (%)"],
            "rows": [
                ["Overall OEE", oee["overall"]],
                ["Availability", oee["availability"]],
                ["Performance", oee["performance"]],
                ["Quality", oee["quality"]],
            ],
        },
        "Equipment Health": {
            "headers": ["Equipment", "Type", "Health Score", "Status"],
            "rows": [[e["name"], e["type"], e["health_score"], e["status"].title()] for e in equipment_health],
        },
    }
    return build_excel_workbook(sheets)


# ─── Executive Summary (combined) ──────────────────────────────────────

def executive_summary_pdf(overview: Dict, oee: Dict, maintenance: Dict, defect_stats: Dict) -> io.BytesIO:
    styles = _pdf_styles()
    story = []

    story.append(Paragraph("Fleet Overview", styles["SectionHeader"]))
    rows = [
        ["Total Equipment", overview["total_equipment"]],
        ["Operational", overview["operational_count"]],
        ["Warning", overview["warning_count"]],
        ["Critical", overview["critical_count"]],
        ["In Maintenance", overview["maintenance_count"]],
        ["Offline", overview["offline_count"]],
        ["Avg Health Score", overview["avg_health_score"]],
    ]
    story.append(_data_table(["Metric", "Value"], rows, col_widths=[3 * inch, 2.5 * inch]))

    story.append(Paragraph("Overall Equipment Effectiveness", styles["SectionHeader"]))
    rows = [
        ["Overall OEE", f"{oee['overall']}%"],
        ["Availability", f"{oee['availability']}%"],
        ["Performance", f"{oee['performance']}%"],
        ["Quality", f"{oee['quality']}%"],
    ]
    story.append(_data_table(["Metric", "Value"], rows, col_widths=[3 * inch, 2.5 * inch]))

    story.append(Paragraph("Maintenance Costs", styles["SectionHeader"]))
    story.append(_data_table(
        ["Metric", "Value"],
        [["Total Cost", f"${maintenance['total_cost']:,.2f}"]],
        col_widths=[3 * inch, 2.5 * inch],
    ))

    story.append(Paragraph("Defect Summary", styles["SectionHeader"]))
    rows = [
        ["Total Defects", defect_stats["total_defects"]],
        ["Critical", defect_stats["critical_count"]],
        ["Resolution Rate", f"{defect_stats['resolution_rate']}%"],
    ]
    story.append(_data_table(["Metric", "Value"], rows, col_widths=[3 * inch, 2.5 * inch]))

    recent = overview.get("recent_alerts", [])[:10]
    if recent:
        story.append(Paragraph("Recent Alerts", styles["SectionHeader"]))
        rows = [[a["created_at"][:10], a["severity"].title(), a["title"]] for a in recent]
        story.append(_data_table(
            ["Date", "Severity", "Title"], rows,
            col_widths=[1 * inch, 1 * inch, 4 * inch],
        ))

    return _build_pdf("Executive Summary Report", _timestamp(), story)